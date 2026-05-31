"""Macro shock sector overlay for model-independent intent and risk adjustments."""

from __future__ import annotations

import logging
import math
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Any

from backtest_shared import FundamentalRow, TradePlan, WorldRegime, clamp

from .config import (
    PROJECT_ROOT,
    SHOCK_OVERLAY_ACTIVE,
    SHOCK_OVERLAY_ALLOW_NEW_INTENTS,
    SHOCK_OVERLAY_BLOCK_LONG_LABELS,
    SHOCK_OVERLAY_BLOCK_SHORT_LABELS,
    SHOCK_OVERLAY_DISABLE_LONG_BOOST_ON_HIGH_LEVERAGE_CREDIT,
    SHOCK_OVERLAY_DISABLE_LONG_BOOST_ON_NEGATIVE_EARNINGS,
    SHOCK_OVERLAY_FULL_SHOCK_SCORE,
    SHOCK_OVERLAY_MAX_INTENT_SCORE_DELTA,
    SHOCK_OVERLAY_MAX_RISK_CUT_PCT,
    SHOCK_OVERLAY_MAX_RISK_UPLIFT_PCT,
    SHOCK_OVERLAY_MIN_SHOCK_SCORE,
    SHOCK_OVERLAY_MODE,
    SHOCK_OVERLAY_POLICY_FILE,
    SHOCK_OVERLAY_SECTOR_BIAS_SHEET,
    SHOCK_OVERLAY_SPECIAL_RULES_SHEET,
    SHOCK_OVERLAY_RISK_OFF_LONG_SLEEVE_ENABLED,
    SHOCK_OVERLAY_RISK_OFF_LONG_SLEEVE_MAX_POSITIONS,
    SHOCK_OVERLAY_RISK_OFF_LONG_SLEEVE_MIN_BIAS,
    SHOCK_OVERLAY_RISK_OFF_LONG_SLEEVE_MIN_INTENT_SCORE,
    SHOCK_OVERLAY_RISK_OFF_LONG_SLEEVE_RISK_MULTIPLIER,
    SHOCK_STRESS_BLOCK_NEGATIVE_BIAS_LONGS,
    SHOCK_STRESS_CREDIT_STRESS_MIN_SCORE,
    SHOCK_STRESS_GUARD_ENABLED,
    SHOCK_STRESS_GUARD_ELEVATED_SCORE,
    SHOCK_STRESS_GUARD_EXTREME_SCORE,
    SHOCK_STRESS_GUARD_HIGH_SCORE,
    SHOCK_STRESS_LONG_RISK_MULTIPLIER_ELEVATED,
    SHOCK_STRESS_LONG_RISK_MULTIPLIER_EXTREME,
    SHOCK_STRESS_LONG_RISK_MULTIPLIER_HIGH,
    SHOCK_STRESS_MAX_LONG_POSITIONS_ELEVATED,
    SHOCK_STRESS_MAX_LONG_POSITIONS_EXTREME,
    SHOCK_STRESS_MAX_LONG_POSITIONS_HIGH,
    SHOCK_STRESS_MAX_POSITIONS_PER_SECTOR_ELEVATED,
    SHOCK_STRESS_MAX_POSITIONS_PER_SECTOR_EXTREME,
    SHOCK_STRESS_MAX_POSITIONS_PER_SECTOR_HIGH,
    SHOCK_STRESS_PORTFOLIO_DAILY_LOSS_LIMIT_PCT,
    SHOCK_STRESS_PORTFOLIO_GUARD_ENABLED,
    SHOCK_STRESS_PORTFOLIO_OPEN_LOSS_LIMIT_PCT,
    SHOCK_STRESS_SECTOR_CAP_ENABLED,
    SHOCK_STRESS_SHORT_MAX_RISK_MULTIPLIER_CREDIT,
)

log = logging.getLogger(__name__)

SHOCK_SCORE_ATTRS = {
    "DEFENSIVE_RISK_OFF": "defensive_risk_off_score",
    "ENERGY_COMMODITY_SHOCK": "energy_commodity_shock_score",
    "RATES_INFLATION_USD_SHOCK": "rates_inflation_usd_shock_score",
    "CREDIT_BANKING_STRESS": "credit_banking_stress_score",
    "POLICY_GEOPOLITICAL_EVENT": "policy_geopolitical_score",
    "METALS_MINING_SHOCK": "metals_mining_shock_score",
}

_SCORE_MODES = {"score_only", "score_and_risk", "full"}
_RISK_MODES = {"risk_only", "score_and_risk", "full"}


@dataclass(frozen=True)
class SpecialRule:
    rule_name: str
    shock_type: str
    sector_key: str
    action: str
    min_score: float
    bias_delta: float = 0.0
    bias_cap: float | None = None
    max_rates_score: float | None = None
    max_credit_score: float | None = None
    direction: str = "ANY"


@dataclass(frozen=True)
class ShockOverlayPolicy:
    sector_bias: dict[tuple[str, str], float]
    special_rules: tuple[SpecialRule, ...]


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower()


def _normalize_sector(value: Any) -> str:
    return " ".join(str(value or "").strip().split()).casefold()


def _normalize_shock_type(value: Any) -> str:
    return str(value or "").strip().upper()


def _normalize_action(value: Any) -> str:
    return str(value or "").strip().lower()


def _normalize_direction(value: Any) -> str:
    direction = str(value or "ANY").strip().upper()
    if direction not in {"ANY", "LONG", "SHORT"}:
        raise ValueError(f"Invalid shock overlay rule direction: {direction!r}")
    return direction


def _float_or_none(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, str) and not value.strip():
        return None
    number = float(value)
    if not math.isfinite(number):
        return None
    return number


def _required_float(row: dict[str, Any], name: str) -> float:
    value = _float_or_none(row.get(name))
    if value is None:
        raise ValueError(f"Shock overlay policy row is missing numeric {name!r}: {row!r}")
    return value


def _iter_sheet_rows(workbook: Any, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Shock overlay workbook is missing sheet {sheet_name!r}")
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_normalize_header(v) for v in rows[0]]
    if not any(headers):
        return []
    result: list[dict[str, Any]] = []
    for raw in rows[1:]:
        if not any(v is not None and str(v).strip() for v in raw):
            continue
        result.append({headers[idx]: value for idx, value in enumerate(raw) if idx < len(headers) and headers[idx]})
    return result


def _policy_path() -> Path:
    path = Path(SHOCK_OVERLAY_POLICY_FILE)
    if not path.is_absolute():
        path = PROJECT_ROOT / path
    return path


@lru_cache(maxsize=1)
def load_shock_overlay_policy() -> ShockOverlayPolicy:
    if not SHOCK_OVERLAY_ACTIVE:
        return ShockOverlayPolicy(sector_bias={}, special_rules=())

    path = _policy_path()
    if not path.exists():
        raise FileNotFoundError(f"Shock overlay policy workbook not found: {path}")

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to load SHOCK_OVERLAY_POLICY_FILE") from exc

    workbook = load_workbook(path, data_only=True, read_only=True)
    sector_bias = _load_sector_bias(workbook)
    special_rules = _load_special_rules(workbook)
    if not sector_bias:
        raise ValueError(f"Shock overlay workbook {path} contains no sector_bias rows")

    log.info(
        "Shock overlay policy loaded mode %s file %s sector bias rows %d special rules %d allow new intents %s",
        SHOCK_OVERLAY_MODE,
        path,
        len(sector_bias),
        len(special_rules),
        SHOCK_OVERLAY_ALLOW_NEW_INTENTS,
    )
    return ShockOverlayPolicy(sector_bias=sector_bias, special_rules=tuple(special_rules))


def _load_sector_bias(workbook: Any) -> dict[tuple[str, str], float]:
    required = {"shock_type", "sector", "bias"}
    sector_bias: dict[tuple[str, str], float] = {}
    for row in _iter_sheet_rows(workbook, SHOCK_OVERLAY_SECTOR_BIAS_SHEET):
        missing = required - set(row)
        if missing:
            raise ValueError(f"sector_bias sheet is missing columns: {', '.join(sorted(missing))}")
        shock_type = _normalize_shock_type(row["shock_type"])
        if shock_type not in SHOCK_SCORE_ATTRS:
            raise ValueError(f"Unknown shock_type in sector_bias sheet: {shock_type!r}")
        sector_key = _normalize_sector(row["sector"])
        if not sector_key:
            raise ValueError(f"sector_bias row has empty sector: {row!r}")
        bias = _required_float(row, "bias")
        if not -1.0 <= bias <= 1.0:
            raise ValueError(f"sector_bias must be between -1 and 1: {row!r}")
        key = (shock_type, sector_key)
        if key in sector_bias:
            raise ValueError(f"Duplicate sector_bias row for shock_type/sector: {shock_type} / {row['sector']}")
        sector_bias[key] = bias
    return sector_bias


def _load_special_rules(workbook: Any) -> list[SpecialRule]:
    if SHOCK_OVERLAY_SPECIAL_RULES_SHEET not in workbook.sheetnames:
        return []

    rules: list[SpecialRule] = []
    required = {"rule_name", "shock_type", "sector", "action", "min_score"}
    for row in _iter_sheet_rows(workbook, SHOCK_OVERLAY_SPECIAL_RULES_SHEET):
        missing = required - set(row)
        if missing:
            raise ValueError(f"special_rules sheet is missing columns: {', '.join(sorted(missing))}")
        rule_name = str(row.get("rule_name") or "").strip()
        if not rule_name:
            raise ValueError(f"special_rules row has empty rule_name: {row!r}")
        shock_type = _normalize_shock_type(row["shock_type"])
        if shock_type not in SHOCK_SCORE_ATTRS:
            raise ValueError(f"Unknown shock_type in special_rules sheet: {shock_type!r}")
        sector_key = _normalize_sector(row["sector"])
        if not sector_key:
            raise ValueError(f"special_rules row has empty sector: {row!r}")
        action = _normalize_action(row["action"])
        if action not in {"boost", "conditional_boost", "penalty", "cap", "floor"}:
            raise ValueError(f"Unknown shock overlay action: {action!r}")
        bias_delta = _float_or_none(row.get("bias_delta")) or 0.0
        bias_cap = _float_or_none(row.get("bias_cap"))
        if bias_cap is not None and not -1.0 <= bias_cap <= 1.0:
            raise ValueError(f"special_rules bias_cap must be between -1 and 1: {row!r}")
        if not -2.0 <= bias_delta <= 2.0:
            raise ValueError(f"special_rules bias_delta must be between -2 and 2: {row!r}")
        rules.append(SpecialRule(
            rule_name=rule_name,
            shock_type=shock_type,
            sector_key=sector_key,
            action=action,
            min_score=_required_float(row, "min_score"),
            bias_delta=bias_delta,
            bias_cap=bias_cap,
            max_rates_score=_float_or_none(row.get("max_rates_score")),
            max_credit_score=_float_or_none(row.get("max_credit_score")),
            direction=_normalize_direction(row.get("direction")),
        ))
    return rules


def _shock_score(regime: WorldRegime, shock_type: str) -> float:
    attr = SHOCK_SCORE_ATTRS[shock_type]
    value = getattr(regime, attr, None)
    if value is None:
        return 0.0
    return float(value)


def _shock_strength(score: float) -> float:
    width = SHOCK_OVERLAY_FULL_SHOCK_SCORE - SHOCK_OVERLAY_MIN_SHOCK_SCORE
    return clamp((float(score) - SHOCK_OVERLAY_MIN_SHOCK_SCORE) / width, 0.0, 1.0)


def shock_stress_score(regime: WorldRegime) -> float:
    value = regime.max_shock_type_score
    if value is not None:
        return float(value)
    return max((_shock_score(regime, shock_type) for shock_type in SHOCK_SCORE_ATTRS), default=0.0)


def _shock_stress_bucket(score: float) -> str:
    if not SHOCK_STRESS_GUARD_ENABLED or score < SHOCK_STRESS_GUARD_ELEVATED_SCORE:
        return ""
    if score >= SHOCK_STRESS_GUARD_EXTREME_SCORE:
        return "extreme"
    if score >= SHOCK_STRESS_GUARD_HIGH_SCORE:
        return "high"
    return "elevated"


def _shock_stress_long_risk_multiplier(score: float) -> float:
    bucket = _shock_stress_bucket(score)
    if bucket == "extreme":
        return SHOCK_STRESS_LONG_RISK_MULTIPLIER_EXTREME
    if bucket == "high":
        return SHOCK_STRESS_LONG_RISK_MULTIPLIER_HIGH
    if bucket == "elevated":
        return SHOCK_STRESS_LONG_RISK_MULTIPLIER_ELEVATED
    return 1.0


def shock_stress_direction_cap(regime: WorldRegime, direction: str, base_cap: int) -> int:
    if direction != "LONG":
        return base_cap
    bucket = _shock_stress_bucket(shock_stress_score(regime))
    if bucket == "extreme":
        return min(base_cap, SHOCK_STRESS_MAX_LONG_POSITIONS_EXTREME)
    if bucket == "high":
        return min(base_cap, SHOCK_STRESS_MAX_LONG_POSITIONS_HIGH)
    if bucket == "elevated":
        return min(base_cap, SHOCK_STRESS_MAX_LONG_POSITIONS_ELEVATED)
    return base_cap


def shock_stress_sector_limit(regime: WorldRegime) -> int | None:
    if not (SHOCK_STRESS_GUARD_ENABLED and SHOCK_STRESS_SECTOR_CAP_ENABLED):
        return None
    bucket = _shock_stress_bucket(shock_stress_score(regime))
    if bucket == "extreme":
        return SHOCK_STRESS_MAX_POSITIONS_PER_SECTOR_EXTREME
    if bucket == "high":
        return SHOCK_STRESS_MAX_POSITIONS_PER_SECTOR_HIGH
    if bucket == "elevated":
        return SHOCK_STRESS_MAX_POSITIONS_PER_SECTOR_ELEVATED
    return None


def shock_stress_plan_block_reason(plan: TradePlan, regime: WorldRegime) -> tuple[str, str] | None:
    score = shock_stress_score(regime)
    if not _shock_stress_bucket(score):
        return None
    if (
        plan.direction == "LONG"
        and SHOCK_STRESS_BLOCK_NEGATIVE_BIAS_LONGS
        and plan.shock_sector_bias < 0.0
    ):
        return (
            "shock_stress_negative_long_bias",
            (
                f"max_shock_type_score {score:.2f} is in stress range and sector shock bias "
                f"{plan.shock_sector_bias:.2f} is negative for long exposure."
            ),
        )
    return None


def shock_stress_portfolio_block_reason(
    regime: WorldRegime,
    account_equity: float,
    day_start_equity: float,
    open_pnl: float,
) -> tuple[str, str] | None:
    if not SHOCK_STRESS_PORTFOLIO_GUARD_ENABLED:
        return None
    score = shock_stress_score(regime)
    if not _shock_stress_bucket(score) or day_start_equity <= 0.0:
        return None
    daily_loss_pct = max(0.0, (day_start_equity - account_equity) / day_start_equity * 100.0)
    if daily_loss_pct >= SHOCK_STRESS_PORTFOLIO_DAILY_LOSS_LIMIT_PCT:
        return (
            "shock_stress_daily_loss_limit",
            (
                f"Portfolio equity is down {daily_loss_pct:.2f}% from day-start equity under "
                f"max_shock_type_score {score:.2f}."
            ),
        )
    open_loss_pct = max(0.0, -open_pnl / day_start_equity * 100.0)
    if open_loss_pct >= SHOCK_STRESS_PORTFOLIO_OPEN_LOSS_LIMIT_PCT:
        return (
            "shock_stress_open_loss_limit",
            (
                f"Open PnL is down {open_loss_pct:.2f}% of day-start equity under "
                f"max_shock_type_score {score:.2f}."
            ),
        )
    return None


def _apply_shock_stress_guard(plan: TradePlan, regime: WorldRegime) -> None:
    score = shock_stress_score(regime)
    if not _shock_stress_bucket(score):
        return

    if plan.direction == "LONG":
        if plan.shock_risk_multiplier > 1.0:
            plan.shock_risk_multiplier = 1.0
        plan.shock_risk_multiplier *= _shock_stress_long_risk_multiplier(score)
    elif _shock_score(regime, "CREDIT_BANKING_STRESS") >= SHOCK_STRESS_CREDIT_STRESS_MIN_SCORE:
        plan.shock_risk_multiplier = min(
            plan.shock_risk_multiplier,
            SHOCK_STRESS_SHORT_MAX_RISK_MULTIPLIER_CREDIT,
        )


def _apply_special_rules(
    bias: float,
    policy: ShockOverlayPolicy,
    regime: WorldRegime,
    sector_key: str,
    direction: str,
) -> float:
    adjusted = bias
    for rule in policy.special_rules:
        if rule.sector_key != sector_key:
            continue
        if rule.direction not in {"ANY", direction}:
            continue
        if _shock_score(regime, rule.shock_type) < rule.min_score:
            continue
        if rule.max_rates_score is not None and _shock_score(regime, "RATES_INFLATION_USD_SHOCK") >= rule.max_rates_score:
            continue
        if rule.max_credit_score is not None and _shock_score(regime, "CREDIT_BANKING_STRESS") >= rule.max_credit_score:
            continue

        if rule.action in {"boost", "conditional_boost", "penalty"}:
            adjusted += rule.bias_delta
            if rule.bias_cap is not None:
                if rule.bias_delta >= 0.0:
                    adjusted = min(adjusted, rule.bias_cap)
                else:
                    adjusted = max(adjusted, rule.bias_cap)
        elif rule.action == "cap" and rule.bias_cap is not None:
            adjusted = min(adjusted, rule.bias_cap)
        elif rule.action == "floor" and rule.bias_cap is not None:
            adjusted = max(adjusted, rule.bias_cap)

    return clamp(adjusted, -1.0, 1.0)


def _apply_guardrails(bias: float, fundamental: FundamentalRow, regime: WorldRegime, direction: str) -> float:
    guarded = bias
    valuation_label = str(fundamental.valuation_label or "").strip().lower()

    if direction == "LONG" and guarded > 0.0:
        if valuation_label in SHOCK_OVERLAY_BLOCK_LONG_LABELS:
            guarded = 0.0
        if (
            guarded > 0.0
            and SHOCK_OVERLAY_DISABLE_LONG_BOOST_ON_HIGH_LEVERAGE_CREDIT
            and fundamental.high_leverage_flag
            and _shock_score(regime, "CREDIT_BANKING_STRESS") >= SHOCK_OVERLAY_MIN_SHOCK_SCORE
        ):
            guarded = 0.0
        if guarded > 0.0 and SHOCK_OVERLAY_DISABLE_LONG_BOOST_ON_NEGATIVE_EARNINGS and fundamental.negative_earnings_flag:
            guarded = 0.0
    elif direction == "SHORT" and guarded < 0.0:
        if valuation_label in SHOCK_OVERLAY_BLOCK_SHORT_LABELS:
            guarded = 0.0

    return guarded


def _copy_regime_to_plan(plan: TradePlan, regime: WorldRegime) -> None:
    plan.dominant_shock_type = regime.dominant_shock_type or ""
    plan.max_shock_type_score = regime.max_shock_type_score
    plan.defensive_risk_off_score = regime.defensive_risk_off_score
    plan.energy_commodity_shock_score = regime.energy_commodity_shock_score
    plan.rates_inflation_usd_shock_score = regime.rates_inflation_usd_shock_score
    plan.credit_banking_stress_score = regime.credit_banking_stress_score
    plan.policy_geopolitical_score = regime.policy_geopolitical_score
    plan.precious_metals_score = regime.precious_metals_score
    plan.industrial_metals_score = regime.industrial_metals_score
    plan.metals_mining_shock_score = regime.metals_mining_shock_score
    plan.metals_mining_subtype = regime.metals_mining_subtype


def shock_sector_bias_for_sector(sector: str, direction: str, regime: WorldRegime) -> float:
    if not SHOCK_OVERLAY_ACTIVE:
        return 0.0

    policy = load_shock_overlay_policy()
    sector_key = _normalize_sector(sector)
    if not sector_key:
        return 0.0

    bias = 0.0
    for shock_type in SHOCK_SCORE_ATTRS:
        strength = _shock_strength(_shock_score(regime, shock_type))
        if strength <= 0.0:
            continue
        bias += strength * policy.sector_bias.get((shock_type, sector_key), 0.0)

    bias = clamp(bias, -1.0, 1.0)
    return _apply_special_rules(bias, policy, regime, sector_key, direction)


def apply_shock_overlay(plan: TradePlan, fundamental: FundamentalRow, regime: WorldRegime) -> None:
    plan.shock_base_intent_score = float(plan.intent_score)
    plan.shock_score_delta = 0.0
    plan.shock_risk_multiplier = 1.0
    plan.shock_sector_bias = 0.0
    _copy_regime_to_plan(plan, regime)

    if not SHOCK_OVERLAY_ACTIVE:
        _apply_shock_stress_guard(plan, regime)
        return

    policy = load_shock_overlay_policy()
    sector_key = _normalize_sector(fundamental.sector)
    if not sector_key:
        _apply_shock_stress_guard(plan, regime)
        return

    bias = 0.0
    for shock_type in SHOCK_SCORE_ATTRS:
        strength = _shock_strength(_shock_score(regime, shock_type))
        if strength <= 0.0:
            continue
        bias += strength * policy.sector_bias.get((shock_type, sector_key), 0.0)

    bias = clamp(bias, -1.0, 1.0)
    bias = _apply_special_rules(bias, policy, regime, sector_key, plan.direction)
    bias = _apply_guardrails(bias, fundamental, regime, plan.direction)
    plan.shock_sector_bias = bias

    directional_bias = bias if plan.direction == "LONG" else -bias
    if SHOCK_OVERLAY_MODE in _SCORE_MODES:
        plan.shock_score_delta = SHOCK_OVERLAY_MAX_INTENT_SCORE_DELTA * directional_bias
        plan.intent_score = float(plan.intent_score) + plan.shock_score_delta

    if SHOCK_OVERLAY_MODE in _RISK_MODES:
        if directional_bias >= 0.0:
            plan.shock_risk_multiplier = 1.0 + directional_bias * SHOCK_OVERLAY_MAX_RISK_UPLIFT_PCT / 100.0
        else:
            plan.shock_risk_multiplier = max(
                0.0,
                1.0 + directional_bias * SHOCK_OVERLAY_MAX_RISK_CUT_PCT / 100.0,
            )
    _apply_shock_stress_guard(plan, regime)


def should_evaluate_disabled_direction(regime_label: str, direction: str) -> bool:
    return (
        SHOCK_OVERLAY_MODE == "full"
        and SHOCK_OVERLAY_RISK_OFF_LONG_SLEEVE_ENABLED
        and regime_label == "RISK-OFF"
        and direction == "LONG"
        and SHOCK_OVERLAY_RISK_OFF_LONG_SLEEVE_MAX_POSITIONS > 0
        and SHOCK_OVERLAY_RISK_OFF_LONG_SLEEVE_RISK_MULTIPLIER > 0.0
    )


def risk_off_long_sleeve_risk(plan: TradePlan, regime_label: str) -> float | None:
    if not should_evaluate_disabled_direction(regime_label, plan.direction):
        return None
    if plan.shock_sector_bias < SHOCK_OVERLAY_RISK_OFF_LONG_SLEEVE_MIN_BIAS:
        return None
    if plan.intent_score < SHOCK_OVERLAY_RISK_OFF_LONG_SLEEVE_MIN_INTENT_SCORE:
        return None
    return SHOCK_OVERLAY_RISK_OFF_LONG_SLEEVE_RISK_MULTIPLIER * plan.shock_risk_multiplier
